import openpyxl as op
import randomdemo as rd

# 一个工作簿(workbook)在创建的时候同时至少也新建了一张工作表(worksheet)
wb = op.Workbook()
# 找到活动的sheet页
sheet = wb.active
# 设置活动的sheet页的名称
sheet.title = 'Game Keys'
# 计算20个随机游戏兑换key写入文件中
keys = rd.key_num(20)
for i in range(21):
    if i == 0:
        sheet["A1"].value = 'GAMEKEYS'
    else:
        sheet["A%d" % (i+1)].value = keys[i-1]
# 最后记得保存
wb.save('gamekeys.xlsx')
# 获取data.xlsx文件
wbr = op.load_workbook('gamekeys.xlsx')
# 看看有哪些sheet页
print(wbr.sheetnames)
# 读取指定的sheet页
sheetr = wbr.get_sheet_by_name('Game Keys')
# 打印出A列中的所有值
for i in sheetr["A"]:
    print(i.value)

